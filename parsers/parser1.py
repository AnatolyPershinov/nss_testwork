import csv
from io import UnsupportedOperation
import os

from typing import List
from dataclasses import dataclass
from wsgiref.util import shift_path_info
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet



@dataclass
class Work:
    index: int
    name: str
    plan_schedule: list
    fact_schedule: list

    def get_dict(self):
        result = {"Название работы": self.name}
        header = [x for x in range(1, len(self.plan_schedule)+1)]
        result.update(dict(zip([str(day) +" plan" for day in header], self.plan_schedule)))
        result.update(dict(zip([str(day) +" fact" for day in header], self.fact_schedule)))
        return result


@dataclass
class Resource:
    index: int
    name: str
    schedule: list
    
    def get_dict(self):
        result = {"Название ресурса": self.name}
        header = [x for x in range(1, len(self.schedule)+1)]
        result.update(dict(zip([str(day) for day in header], self.schedule)))
        return result


def remove_extra_char(s: str) -> str:
    # удаление из строки символов переноса, пробелов, заглавных букв
    if type(s) is str:
        return s.replace("\n", "").replace(" ", "").lower()
    else:
        return ""


def parser(sheet: Worksheet, days = int) ->  tuple[list[Work], list[Resource]]:
    # метод собирает информацию об активностях и ресурсах
    works: List[Work] = []
    resources: List[Resource] = []
    # поиск заголовка
    header_row = 0
    for cell in sheet["A"]:
        if cell.value is None: 
            continue
        elif remove_extra_char(cell.value) == "№п/п":
            header_row = cell.row
            break
    # парсинг заголовка
    # поиск колонки с именем, поиск колонки-разделителя
    name_column = 0
    ident_column = 0

    for cell in sheet[header_row]:
        if cell.value is None:
            continue
        elif remove_extra_char(cell.value) == "наименованиеработ":
            name_column = cell.column
        elif remove_extra_char(cell.value) == "днимес.":
            ident_column = cell.column # индекс колонки план/факт

    # часть 1. парсинг активностей
    # поиск колонки с идентификацией фактический и плановой активности
    # информация об активности содержится в строках с записью "план"
    for cell in sheet[get_column_letter(ident_column)]:
        if cell.value != "план":
            continue

        row = sheet[cell.row]
        fact_row = sheet[cell.row+1] # строка с фактической активностью

        # списки с календарями активности
        # информация об активнсти по дням расположена слева от колокни "план/факт"
        # поэтому берертся срез списка, [индекс колонки:индекс колокни + кол-во дней в месяце] 
        plan_schedule = [c.value for c in row[ident_column:ident_column+days]]
        fact_schedule = [c.value for c in fact_row[ident_column:ident_column+days]]
        try:
            works.append(Work(
            index=len(works),
            name=row[1].value+"_act",
            plan_schedule=[0 if v is None else v for v in plan_schedule], # заменить None на 0
            fact_schedule=[0 if v is None else v for v in fact_schedule]
        ))
        except Exception as e:
            if row[1].value is None:
                print(f"unknown action name in sheet: {sheet.title} row: {cell.row}. skip it.")
            else:
                print(Exception)                
        
    # часть 2. поиск ресурсов
    start_pos = 0
    end_pos = 0
    resource_rows = []

    for cell in sheet[get_column_letter(name_column)]:
        if cell.value is None:
            continue
        elif remove_extra_char(cell.value) == "наименованиеимаркатехники(механизма),оборудования":
            start_pos = cell.row + 2
        elif remove_extra_char(cell.value) == "наименованиесубподряднойорганизации":
            end_pos = cell.row-3
            break
    
    for cell in sheet[get_column_letter(name_column)][start_pos:end_pos]:
        if cell.value is not None and remove_extra_char(cell.value) != "наименованиедолжностей,профессий":
            resource_rows.append(sheet[cell.row])

    for cell in sheet[get_column_letter(name_column)][end_pos+6:]:
        if cell.value is not None and remove_extra_char(cell.value) != "наименованиесубподряднойорганизации" and cell.value != 2:
            if sheet[cell.row][0].value is None:
                continue
            resource_rows.append(sheet[cell.row])

    for resource in resource_rows:
        plan_schedule = [c.value for c in resource[ident_column:ident_column+days]]
        
        resources.append(Resource(
            index=len(resources),
            name=resource[1].value+"_res", 
            schedule=[0 if v is None else v for v in plan_schedule]
        ))

    return works, resources


def save_to_csv(filename, array):
    with open(filename, "w", encoding="UTF-8", newline='\r\n') as csvfile:
        header = array[0].get_dict().keys()
        writer = csv.DictWriter(f=csvfile, fieldnames=header, lineterminator="\n")
        writer.writeheader()
        for k in array:
            writer.writerow(k.get_dict())


def main(path: str):
    workbook = load_workbook(path)

    months = [
        ('февраль', 28),('март',31), ('апрель', 30), 
        ('май', 31), ('июнь',30), ('июль', 31), 
        ('август', 31), ('сентябрь',30), ('ОКТЯБРЬ', 31), 
        ('Ноябрь', 30), ('Декабрь', 31)
    ]

    for month in months:
        ws = workbook.get_sheet_by_name(month[0])
        works, resources = parser(sheet=ws, days=month[1])
        path = f"results/file1/{month[0].lower()}/"

        if not os.path.exists(path):
            os.makedirs(path)

        save_to_csv(path+"activities.csv", works)
        save_to_csv(path+"resources.csv", resources)
